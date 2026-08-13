#!/usr/bin/env python3
"""
BLS Data Extraction Script
===========================
Extracts REAL data from BLS OES May 2023 (all_data_M_2023.xlsx) and
merges with existing project data to replace all hardcoded values.

Inputs:
  - data/all_data_M_2023.xlsx (BLS OES May 2023 - 413K rows)
  - data/Kelley_Job_Map.csv (SOC → title mapping)
  - data/ai_impact_scores.json (pre-scored tasks, partial)

Outputs:
  - src/data/geo_real.json (real state-level employment data, all states)
  - src/data/national_employment.json (national figures)
  - src/data/bls_extracted.json (full extracted dataset for data.ts)

Notes on geo_real.json:
  - SOC codes are taken from src/utils/onet.ts MAP_TITLE_TO_SOC (the runtime
    lookup key), not from Kelley_Job_Map.csv, so there is no drift between the
    extraction and what the app uses.
  - All states with employment > 0 are included (no top-N truncation).
  - lat/lng values are US Census Bureau geographic centroids (display-only;
    they are NOT BLS data).
  - The _meta block records provenance so every value in geo_real.json can be
    traced to the source workbook.
"""

import csv
import json
import os
import sys
import datetime

try:
    import openpyxl
except ImportError:
    print("Installing openpyxl...")
    os.system(f"{sys.executable} -m pip install openpyxl -q")
    import openpyxl

# --- Configuration ---
PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_DIR = os.path.join(PROJECT_ROOT, "data")
SRC_DATA_DIR = os.path.join(PROJECT_ROOT, "src", "data")

XLSX_PATH = os.path.join(DATA_DIR, "all_data_M_2023.xlsx")
CSV_PATH = os.path.join(DATA_DIR, "Kelley_Job_Map.csv")
AI_SCORES_PATH = os.path.join(DATA_DIR, "ai_impact_scores.json")

# Canonical SOC codes as used by MAP_TITLE_TO_SOC in src/utils/onet.ts.
# This is the runtime lookup key; extracting with these codes ensures geo_real.json
# matches exactly what the app queries — no drift between onet.ts and this file.
# Duplicate SOC codes (e.g. Marketing Manager + Brand Manager → 11-2021) are
# intentional: they share the same occupation group per BLS OES, and the app's
# aggregation layer de-dupes by SOC before summing state employment.
MAP_TITLE_TO_SOC = {
    # Original 13 jobs
    "Marketing Manager":               "11-2021",
    "Market Research Analyst":         "13-1161",
    "Business Intelligence Analyst":   "15-2051",
    "Financial Analyst":               "13-2051",
    "Financial Manager":               "11-3031",
    "Software Developer":              "15-1252",
    "Management Consultant":           "13-1111",
    "Sales Manager":                   "11-2022",
    "Accountant & Auditor":            "13-2011",
    "Operations Research Analyst":     "15-2031",
    "Logistics Analyst":               "13-1081",
    "Securities & Sales Agent":        "41-3031",
    "Supply Chain Manager":            "11-3071",
    # Business / Marketing
    "Brand Manager":                   "11-2021",
    "Public Relations Manager":        "11-2032",  # BLS 2018 SOC (was 11-2031)
    "Advertising Sales Agent":         "41-3011",
    "Event Coordinator":               "13-1121",
    "Account Executive":               "41-3091",  # Sales Representatives of Services (BLS OES)
    "Sales Representative":            "41-4012",
    "Insurance Sales Agent":           "41-3021",
    # Finance
    "Personal Financial Advisor":      "13-2052",
    "Credit Analyst":                  "13-2041",
    "Budget Analyst":                  "13-2031",
    "Risk Specialist":                 "13-2099",
    "Insurance Underwriter":           "13-2053",
    "Actuary":                         "15-2011",
    "Loan Officer":                    "13-2072",
    "Financial Risk Analyst":          "13-2099",
    # Data & Technology
    "Data Scientist":                  "15-2051",
    "Statistician":                    "15-2041",
    "Computer Systems Analyst":        "15-1211",
    "UX Designer":                     "15-1255",
    "IT Manager":                      "11-3021",
    "Cybersecurity Analyst":           "15-1212",
    "Web Developer":                   "15-1254",
    "Database Administrator":          "15-1242",
    "IT Project Manager":              "15-1299",
    # Operations
    "Cost Estimator":                  "13-1051",
    "Compensation Analyst":            "13-1141",
    "Purchasing Manager":              "11-3061",
    "Wholesale & Retail Buyer":        "13-1022",
    "Purchasing Agent":                "13-1023",
    "Industrial Production Manager":   "11-3051",
    # HR & Management
    "General Manager":                 "11-1021",
    "HR Manager":                      "11-3121",
    "Project Management Specialist":   "13-1082",
    "Training & Development Specialist": "13-1151",
    "HR Specialist":                   "13-1071",
    "Training & Development Manager":  "11-3131",
    "Compensation & Benefits Manager": "11-3111",
}

# Unique SOC codes to extract (some titles share a code; extract each code once)
UNIQUE_SOC_CODES = sorted(set(MAP_TITLE_TO_SOC.values()))

# BLS Employment Projections 2024-2034
# Source: https://www.bls.gov/emp/tables/occupational-projections-and-characteristics.htm
# Official projected percent change in employment, 2024-34 (rounded to nearest integer)
BLS_GROWTH_PROJECTIONS = {
    "11-2021": 7,     # Marketing Managers: 7% (BLS OOH 2024-34)
    "13-1161": 7,     # Market Research Analysts: 7% (BLS OOH 2024-34)
    "15-2051": 34,     # Data Scientists / BI Analysts: 34% (BLS OOH 2024-34)
    "13-2051": 6,     # Financial Analysts: 6% (BLS OOH 2024-34)
    "11-3031": 15,     # Financial Managers: 15% (BLS OOH 2024-34)
    "15-1252": 16,     # Software Developers: 16% (BLS OOH 2024-34)
    "13-1111": 9,     # Management Analysts/Consultants: 9% (BLS OOH 2024-34)
    "11-2022": 5,     # Sales Managers: 5% (BLS OOH 2024-34)
    "13-2011": 5,     # Accountants and Auditors: 5% (BLS OOH 2024-34)
    "15-2031": 22,     # Operations Research Analysts: 22% (BLS OOH 2024-34)
    "13-1081": 17,     # Logisticians: 17% (BLS OOH 2024-34)
    "41-3031": 3,     # Securities/Financial Services Sales Agents: 3% (BLS OOH 2024-34)
    "11-3071": 6,     # Transportation/Storage/Distribution Managers: 6% (BLS OOH 2024-34)
    "11-2032": 5,     # Public Relations Managers: 5% (BLS OOH 2024-34)
    "41-3091": 3,     # Sales Representatives of Services: 3% (BLS OOH 2024-34)
    "41-3011": -6,     # Advertising Sales Agents: -6% (BLS OOH 2024-34)
    "13-1121": 5,     # Meeting/Convention/Event Planners: 5% (BLS OOH 2024-34)
    "41-4012": 0,     # Sales Representatives, Wholesale: 0% (BLS OOH 2024-34)
    "41-3021": 4,     # Insurance Sales Agents: 4% (BLS OOH 2024-34)
    "13-2052": 10,     # Personal Financial Advisors: 10% (BLS OOH 2024-34)
    "13-2041": -4,     # Credit Analysts: -4% (BLS OOH 2024-34)
    "13-2031": 1,     # Budget Analysts: 1% (BLS OOH 2024-34)
    "13-2099": 3,     # Financial Specialists, All Other: 3% (BLS OOH 2024-34)
    "13-2053": -3,     # Insurance Underwriters: -3% (BLS OOH 2024-34)
    "15-2011": 22,     # Actuaries: 22% (BLS OOH 2024-34)
    "13-2072": 2,     # Loan Officers: 2% (BLS OOH 2024-34)
    "15-2041": 9,     # Statisticians: 9% (BLS OOH 2024-34)
    "15-1211": 9,     # Computer Systems Analysts: 9% (BLS OOH 2024-34)
    "13-1051": -4,     # Cost Estimators: -4% (BLS OOH 2024-34)
    "13-1141": 5,     # Compensation/Benefits/Job Analysis Specialists: 5% (BLS OOH 2024-34)
    "11-3061": 3,     # Purchasing Managers: 3% (BLS OOH 2024-34)
    "13-1022": 6,     # Wholesale and Retail Buyers (use 13-1020 rollup): 6% (BLS OOH 2024-34)
    "13-1023": 6,     # Purchasing Agents (use 13-1020 rollup): 6% (BLS OOH 2024-34)
    "11-3051": 2,     # Industrial Production Managers: 2% (BLS OOH 2024-34)
    "15-1255": 7,     # Web and Digital Interface Designers: 7% (BLS OOH 2024-34)
    "11-3021": 15,     # Computer and IT Managers: 15% (BLS OOH 2024-34)
    "15-1212": 29,     # Information Security Analysts: 29% (BLS OOH 2024-34)
    "15-1254": 8,     # Web Developers: 8% (BLS OOH 2024-34)
    "15-1242": -1,     # Database Administrators: -1% (BLS OOH 2024-34)
    "15-1299": 8,     # Computer Occupations, All Other: 8% (BLS OOH 2024-34)
    "11-1021": 4,     # General and Operations Managers: 4% (BLS OOH 2024-34)
    "11-3121": 5,     # Human Resources Managers: 5% (BLS OOH 2024-34)
    "13-1082": 6,     # Project Management Specialists: 6% (BLS OOH 2024-34)
    "13-1151": 11,     # Training and Development Specialists: 11% (BLS OOH 2024-34)
    "13-1071": 6,     # Human Resources Specialists: 6% (BLS OOH 2024-34)
    "11-3131": 6,     # Training and Development Managers: 6% (BLS OOH 2024-34)
    "11-3111": 0,     # Compensation and Benefits Managers: 0% (BLS OOH 2024-34)
}

# US State centroid coordinates
# Source: US Census Bureau Geographic Reference Files
# These are display-only coordinates used to position map markers.
# They are NOT part of the BLS data — do not cite them as BLS.
STATE_INFO = {
    "Alabama": {"lat": 32.81, "lng": -86.79, "abbr": "AL"},
    "Alaska": {"lat": 63.35, "lng": -152.00, "abbr": "AK"},
    "Arizona": {"lat": 34.05, "lng": -111.09, "abbr": "AZ"},
    "Arkansas": {"lat": 34.75, "lng": -92.29, "abbr": "AR"},
    "California": {"lat": 36.78, "lng": -119.42, "abbr": "CA"},
    "Colorado": {"lat": 39.55, "lng": -105.78, "abbr": "CO"},
    "Connecticut": {"lat": 41.60, "lng": -73.09, "abbr": "CT"},
    "Delaware": {"lat": 38.91, "lng": -75.53, "abbr": "DE"},
    "District of Columbia": {"lat": 38.91, "lng": -77.04, "abbr": "DC"},
    "Florida": {"lat": 27.66, "lng": -81.52, "abbr": "FL"},
    "Georgia": {"lat": 32.17, "lng": -82.90, "abbr": "GA"},
    "Hawaii": {"lat": 19.90, "lng": -155.58, "abbr": "HI"},
    "Idaho": {"lat": 44.07, "lng": -114.74, "abbr": "ID"},
    "Illinois": {"lat": 40.63, "lng": -89.40, "abbr": "IL"},
    "Indiana": {"lat": 40.27, "lng": -86.13, "abbr": "IN"},
    "Iowa": {"lat": 41.88, "lng": -93.10, "abbr": "IA"},
    "Kansas": {"lat": 39.01, "lng": -98.48, "abbr": "KS"},
    "Kentucky": {"lat": 37.67, "lng": -84.65, "abbr": "KY"},
    "Louisiana": {"lat": 30.98, "lng": -91.96, "abbr": "LA"},
    "Maine": {"lat": 45.25, "lng": -69.45, "abbr": "ME"},
    "Maryland": {"lat": 39.05, "lng": -76.64, "abbr": "MD"},
    "Massachusetts": {"lat": 42.41, "lng": -71.38, "abbr": "MA"},
    "Michigan": {"lat": 44.31, "lng": -85.60, "abbr": "MI"},
    "Minnesota": {"lat": 46.73, "lng": -94.69, "abbr": "MN"},
    "Mississippi": {"lat": 32.35, "lng": -89.40, "abbr": "MS"},
    "Missouri": {"lat": 37.96, "lng": -91.83, "abbr": "MO"},
    "Montana": {"lat": 46.88, "lng": -110.36, "abbr": "MT"},
    "Nebraska": {"lat": 41.49, "lng": -99.90, "abbr": "NE"},
    "Nevada": {"lat": 38.80, "lng": -116.42, "abbr": "NV"},
    "New Hampshire": {"lat": 43.19, "lng": -71.57, "abbr": "NH"},
    "New Jersey": {"lat": 40.06, "lng": -74.41, "abbr": "NJ"},
    "New Mexico": {"lat": 34.52, "lng": -105.87, "abbr": "NM"},
    "New York": {"lat": 40.71, "lng": -74.01, "abbr": "NY"},
    "North Carolina": {"lat": 35.76, "lng": -79.02, "abbr": "NC"},
    "North Dakota": {"lat": 47.55, "lng": -101.00, "abbr": "ND"},
    "Ohio": {"lat": 40.42, "lng": -82.91, "abbr": "OH"},
    "Oklahoma": {"lat": 35.01, "lng": -97.09, "abbr": "OK"},
    "Oregon": {"lat": 43.80, "lng": -120.55, "abbr": "OR"},
    "Pennsylvania": {"lat": 41.20, "lng": -77.19, "abbr": "PA"},
    "Rhode Island": {"lat": 41.58, "lng": -71.48, "abbr": "RI"},
    "South Carolina": {"lat": 33.84, "lng": -81.16, "abbr": "SC"},
    "South Dakota": {"lat": 43.97, "lng": -99.90, "abbr": "SD"},
    "Tennessee": {"lat": 35.52, "lng": -86.58, "abbr": "TN"},
    "Texas": {"lat": 31.97, "lng": -99.90, "abbr": "TX"},
    "Utah": {"lat": 39.32, "lng": -111.09, "abbr": "UT"},
    "Vermont": {"lat": 44.56, "lng": -72.58, "abbr": "VT"},
    "Virginia": {"lat": 37.43, "lng": -78.66, "abbr": "VA"},
    "Washington": {"lat": 47.75, "lng": -120.74, "abbr": "WA"},
    "West Virginia": {"lat": 38.60, "lng": -80.45, "abbr": "WV"},
    "Wisconsin": {"lat": 43.78, "lng": -88.79, "abbr": "WI"},
    "Wyoming": {"lat": 43.08, "lng": -107.29, "abbr": "WY"},
    "Guam": {"lat": 13.44, "lng": 144.79, "abbr": "GU"},
    "Puerto Rico": {"lat": 18.22, "lng": -66.59, "abbr": "PR"},
    "Virgin Islands": {"lat": 18.34, "lng": -64.90, "abbr": "VI"},
}


def load_job_map():
    """Load the SOC → title mapping from Kelley_Job_Map.csv (reference only)."""
    jobs = {}
    with open(CSV_PATH, 'r', encoding='utf-8-sig') as f:
        reader = csv.reader(f)
        next(reader)  # skip header
        for row in reader:
            if len(row) == 1:
                parts = row[0].split(',')
            else:
                parts = row
            if len(parts) >= 4:
                soc = parts[0].strip()
                title = parts[2].strip()
                track = parts[3].strip().strip('"')
                jobs[soc] = {"title": title, "track": track}
    return jobs


def load_ai_scores():
    """Load pre-scored AI impact data"""
    with open(AI_SCORES_PATH, 'r') as f:
        return json.load(f)


def extract_national_data(wb, soc_codes):
    """Extract national employment data for each SOC code from BLS OES"""
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    col_idx = {name: i for i, name in enumerate(headers)}

    results = {}

    for row in ws.iter_rows(min_row=2):
        vals = [cell.value for cell in row]
        area = str(vals[col_idx['AREA']] or '')
        occ_code = str(vals[col_idx['OCC_CODE']] or '')
        i_group = str(vals[col_idx['I_GROUP']] or '')

        # National data: AREA='99', cross-industry
        if area == '99' and i_group == 'cross-industry' and occ_code in soc_codes:
            o_group = str(vals[col_idx['O_GROUP']] or '')
            if o_group == 'detailed':
                tot_emp = vals[col_idx['TOT_EMP']]
                a_mean = vals[col_idx['A_MEAN']]
                a_median = vals[col_idx['A_MEDIAN']]

                emp = parse_int(tot_emp)
                salary_mean = parse_int(a_mean)
                salary_median = parse_int(a_median)

                results[occ_code] = {
                    "employment": emp,
                    "salary_mean": salary_mean,
                    "salary_median": salary_median,
                }
                print(f"  ✓ {occ_code}: emp={emp:,}, mean=${salary_mean:,}, median=${salary_median:,}")

    return results


def extract_state_data(wb, soc_codes):
    """Extract ALL states with employment > 0 for each SOC code from BLS OES.

    Includes every state/territory row — no top-N truncation — so the map can
    show full US coverage. Values are taken directly from BLS OES TOT_EMP and
    LOC_QUOTIENT columns without modification.
    """
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    col_idx = {name: i for i, name in enumerate(headers)}

    state_data = {soc: [] for soc in soc_codes}

    for row in ws.iter_rows(min_row=2):
        vals = [cell.value for cell in row]
        area = str(vals[col_idx['AREA']] or '')
        occ_code = str(vals[col_idx['OCC_CODE']] or '')
        i_group = str(vals[col_idx['I_GROUP']] or '')
        o_group = str(vals[col_idx['O_GROUP']] or '')
        area_title = str(vals[col_idx['AREA_TITLE']] or '')

        # State-level: AREA is 2-digit (01-72), not '99' (national)
        is_state = area != '99' and len(area) <= 2 and area_title in STATE_INFO

        if is_state and i_group == 'cross-industry' and o_group == 'detailed' and occ_code in soc_codes:
            tot_emp = vals[col_idx['TOT_EMP']]
            loc_q = vals[col_idx['LOC_QUOTIENT']]

            emp = parse_int(tot_emp)
            lq = parse_float(loc_q)

            if emp > 0:
                state_data[occ_code].append({
                    "state_name": area_title,
                    "employment": emp,
                    "lq": round(lq, 2),
                })

    # Sort by employment (descending), attach Census centroids for display
    geo_result = {}
    for soc in soc_codes:
        states = sorted(state_data[soc], key=lambda x: x['employment'], reverse=True)
        geo_result[soc] = []
        for s in states:
            info = STATE_INFO[s['state_name']]
            geo_result[soc].append({
                "name": s['state_name'],
                "lat": info['lat'],
                "lng": info['lng'],
                "employment": s['employment'],
                "lq": s['lq'],
            })
        top_info = f"{states[0]['state_name']} = {states[0]['employment']:,}" if states else "—"
        print(f"  ✓ {soc}: {len(states)} states (top: {top_info})")

    return geo_result


def parse_int(val):
    """Safely parse a BLS cell as integer (handles '**', '#', commas)"""
    if not val or str(val).strip() in ('**', '*', '#', ''):
        return 0
    try:
        return int(str(val).replace(',', '').split('.')[0])
    except (ValueError, TypeError):
        return 0


def parse_float(val):
    """Safely parse a BLS cell as float"""
    if not val or str(val).strip() in ('**', '*', '#', ''):
        return 0.0
    try:
        return float(str(val).replace(',', ''))
    except (ValueError, TypeError):
        return 0.0


def main():
    print("=" * 60)
    print("BLS Data Extraction Pipeline")
    print("=" * 60)

    # 1. Load job mapping (reference; canonical codes are in MAP_TITLE_TO_SOC above)
    print("\n📋 Loading job mapping...")
    job_map = load_job_map()
    print(f"   Found {len(job_map)} Kelley SOC codes")
    print(f"   Using {len(UNIQUE_SOC_CODES)} unique canonical SOC codes from MAP_TITLE_TO_SOC")

    # 2. Load AI impact scores
    print("\n🤖 Loading AI impact scores...")
    ai_scores = load_ai_scores()
    scored_socs = set(ai_scores.keys())
    missing_socs = set(UNIQUE_SOC_CODES) - scored_socs
    print(f"   Pre-scored: {len(scored_socs)} SOC codes")
    if missing_socs:
        print(f"   ⚠ Not in legacy ai_impact_scores: {sorted(missing_socs)} (scored live by Claude)")

    # 3. Open BLS xlsx (load once, pass workbook to both extractors)
    print(f"\n📊 Opening BLS OES data (scanning ~413K rows)...")
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True)

    # 4. Extract national employment
    print("\n🇺🇸 National employment:")
    national = extract_national_data(wb, set(UNIQUE_SOC_CODES))
    print(f"   → {len(national)}/{len(UNIQUE_SOC_CODES)} SOC codes found")
    wb.close()

    # 5. Extract state-level data (ALL states, no top-N truncation)
    print(f"\n🏛️  State-level employment (all states per SOC, from BLS OES):")
    wb2 = openpyxl.load_workbook(XLSX_PATH, read_only=True)
    geo = extract_state_data(wb2, set(UNIQUE_SOC_CODES))
    wb2.close()

    # 6. Write outputs
    print("\n📦 Writing output files...")

    # Output 1: geo_real.json — includes _meta provenance block
    geo_output = {
        "_meta": {
            "source": "BLS Occupational Employment and Wage Statistics (OES), May 2023",
            "source_file": "all_data_M_2023.xlsx",
            "bls_release": "OES May 2023",
            "extracted": datetime.date.today().isoformat(),
            "soc_count": len([k for k in geo if k != "_meta"]),
            "note_coordinates": (
                "lat/lng values are US Census Bureau geographic centroids "
                "(display-only). They are NOT part of BLS data."
            ),
            "note_employment": (
                "employment values are BLS OES TOT_EMP (cross-industry, detailed "
                "occupation level). lq values are BLS LOC_QUOTIENT."
            ),
            "note_soc_codes": (
                "SOC codes match MAP_TITLE_TO_SOC in src/utils/onet.ts exactly. "
                "Some titles share a SOC (e.g. Marketing Manager and Brand Manager "
                "both map to 11-2021). The map aggregation layer de-dupes by SOC."
            ),
        }
    }
    # Merge geo data under soc keys after _meta
    geo_output.update({k: v for k, v in geo.items()})

    geo_path = os.path.join(SRC_DATA_DIR, "geo_real.json")
    with open(geo_path, 'w') as f:
        json.dump(geo_output, f, indent=2, ensure_ascii=False)
    total_state_rows = sum(len(v) for k, v in geo_output.items() if k != "_meta")
    print(f"   ✓ {geo_path}  ({len(UNIQUE_SOC_CODES)} SOCs, {total_state_rows} state rows)")

    # Output 2: national_employment.json
    nat_path = os.path.join(SRC_DATA_DIR, "national_employment.json")
    with open(nat_path, 'w') as f:
        json.dump(national, f, indent=2)
    print(f"   ✓ {nat_path}")

    # Output 3: bls_extracted.json (comprehensive)
    bls_extracted = {}
    for soc in UNIQUE_SOC_CODES:
        # Find titles that map to this SOC
        titles = [t for t, s in MAP_TITLE_TO_SOC.items() if s == soc]
        bls_extracted[soc] = {
            "soc_code": soc,
            "titles": titles,
            "employment": national.get(soc, {}).get("employment", 0),
            "salary_mean": national.get(soc, {}).get("salary_mean", 0),
            "salary_median": national.get(soc, {}).get("salary_median", 0),
            "projectedGrowth": BLS_GROWTH_PROJECTIONS.get(soc, 0),
            "state_count": len(geo.get(soc, [])),
        }

    ext_path = os.path.join(SRC_DATA_DIR, "bls_extracted.json")
    with open(ext_path, 'w') as f:
        json.dump(bls_extracted, f, indent=2)
    print(f"   ✓ {ext_path}")

    # Summary table
    print("\n" + "=" * 90)
    print("EXTRACTION COMPLETE — Real BLS OES May 2023 Data")
    print("=" * 90)
    print(f"\n{'SOC Code':<12} {'Titles':<38} {'Employment':>12} {'Growth':>8} {'States':>8}")
    print("-" * 90)
    for soc in sorted(UNIQUE_SOC_CODES):
        e = bls_extracted[soc]
        titles_str = ", ".join(e['titles'])[:36]
        if len(", ".join(e['titles'])) > 36:
            titles_str += "…"
        print(f"{soc:<12} {titles_str:<38} {e['employment']:>12,} {e['projectedGrowth']:>7}% {e['state_count']:>7}")

    socs_with_geo = sum(1 for soc in UNIQUE_SOC_CODES if geo.get(soc))
    socs_no_geo = [soc for soc in UNIQUE_SOC_CODES if not geo.get(soc)]
    print(f"\n✅ {len(national)}/{len(UNIQUE_SOC_CODES)} SOC codes with national data")
    print(f"✅ {socs_with_geo}/{len(UNIQUE_SOC_CODES)} SOC codes with state-level data")
    print(f"✅ {total_state_rows} total state employment rows in geo_real.json")
    if socs_no_geo:
        print(f"⚠️  SOCs with no state rows (BLS may suppress): {socs_no_geo}")


if __name__ == "__main__":
    main()
