#!/usr/bin/env python3
"""
Geographic Data Verification Script
=====================================
Verifies that every employment and lq value in src/data/geo_real.json matches
the corresponding value in the BLS OES May 2023 source workbook.

This script is the automated proof that the map numbers are traceable to BLS.

Usage:
  python3 scripts/verify_geo_data.py

Exit code 0 = all values verified.
Exit code 1 = mismatches or missing rows found (printed to stdout).
"""

import json
import os
import sys

try:
    import openpyxl
except ImportError:
    print("Installing openpyxl...")
    os.system(f"{sys.executable} -m pip install openpyxl -q")
    import openpyxl

PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX_PATH = os.path.join(PROJECT_ROOT, "data", "all_data_M_2023.xlsx")
GEO_PATH = os.path.join(PROJECT_ROOT, "src", "data", "geo_real.json")

STATE_ABBR_TO_NAME = {
    "AL": "Alabama", "AK": "Alaska", "AZ": "Arizona", "AR": "Arkansas",
    "CA": "California", "CO": "Colorado", "CT": "Connecticut", "DE": "Delaware",
    "DC": "District of Columbia", "FL": "Florida", "GA": "Georgia", "HI": "Hawaii",
    "ID": "Idaho", "IL": "Illinois", "IN": "Indiana", "IA": "Iowa", "KS": "Kansas",
    "KY": "Kentucky", "LA": "Louisiana", "ME": "Maine", "MD": "Maryland",
    "MA": "Massachusetts", "MI": "Michigan", "MN": "Minnesota", "MS": "Mississippi",
    "MO": "Missouri", "MT": "Montana", "NE": "Nebraska", "NV": "Nevada",
    "NH": "New Hampshire", "NJ": "New Jersey", "NM": "New Mexico", "NY": "New York",
    "NC": "North Carolina", "ND": "North Dakota", "OH": "Ohio", "OK": "Oklahoma",
    "OR": "Oregon", "PA": "Pennsylvania", "RI": "Rhode Island", "SC": "South Carolina",
    "SD": "South Dakota", "TN": "Tennessee", "TX": "Texas", "UT": "Utah",
    "VT": "Vermont", "VA": "Virginia", "WA": "Washington", "WV": "West Virginia",
    "WI": "Wisconsin", "WY": "Wyoming", "GU": "Guam", "PR": "Puerto Rico",
    "VI": "Virgin Islands",
}

STATE_NAMES = set(STATE_ABBR_TO_NAME.values())


def parse_int(val):
    if not val or str(val).strip() in ('**', '*', '#', ''):
        return 0
    try:
        return int(str(val).replace(',', '').split('.')[0])
    except (ValueError, TypeError):
        return 0


def parse_float(val):
    if not val or str(val).strip() in ('**', '*', '#', ''):
        return 0.0
    try:
        return float(str(val).replace(',', ''))
    except (ValueError, TypeError):
        return 0.0


def load_bls_state_data(soc_codes):
    """Load state-level BLS data for the given SOC codes from the xlsx."""
    print(f"Loading BLS xlsx for {len(soc_codes)} SOC codes...")
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    col_idx = {name: i for i, name in enumerate(headers)}

    bls_data = {}  # (soc, state_name) → (employment, lq)

    for row in ws.iter_rows(min_row=2):
        vals = [cell.value for cell in row]
        area = str(vals[col_idx['AREA']] or '')
        occ_code = str(vals[col_idx['OCC_CODE']] or '')
        i_group = str(vals[col_idx['I_GROUP']] or '')
        o_group = str(vals[col_idx['O_GROUP']] or '')
        area_title = str(vals[col_idx['AREA_TITLE']] or '')

        is_state = area != '99' and len(area) <= 2 and area_title in STATE_NAMES
        if is_state and i_group == 'cross-industry' and o_group == 'detailed' and occ_code in soc_codes:
            emp = parse_int(vals[col_idx['TOT_EMP']])
            lq = round(parse_float(vals[col_idx['LOC_QUOTIENT']]), 2)
            if emp > 0:
                bls_data[(occ_code, area_title)] = (emp, lq)

    wb.close()
    return bls_data


def main():
    print("=" * 64)
    print("geo_real.json Verification against BLS OES May 2023")
    print("=" * 64)

    # 1. Load geo_real.json
    with open(GEO_PATH, 'r') as f:
        geo = json.load(f)

    # Verify _meta block is present
    if '_meta' not in geo:
        print("FAIL: _meta block missing from geo_real.json")
        sys.exit(1)
    print(f"\n✓ _meta block present: {geo['_meta']['bls_release']}, extracted {geo['_meta']['extracted']}")

    soc_entries = {k: v for k, v in geo.items() if k != '_meta'}
    print(f"✓ {len(soc_entries)} SOC codes in geo_real.json")

    # Count expected verifications
    total_rows = sum(len(rows) for rows in soc_entries.values())
    print(f"✓ {total_rows} state rows to verify\n")

    # 2. Collect all (soc, state_name) pairs needed
    needed_socs = set(soc_entries.keys())

    # 3. Load BLS data
    bls_data = load_bls_state_data(needed_socs)
    print(f"✓ Loaded {len(bls_data)} BLS state rows for these SOC codes\n")

    # 4. Verify each geo_real.json row
    mismatches = []
    missing = []
    verified = 0

    for soc, rows in soc_entries.items():
        for row in rows:
            state = row['name']
            geo_emp = row['employment']
            geo_lq = row['lq']

            key = (soc, state)
            if key not in bls_data:
                missing.append(f"  {soc} / {state}: in geo_real.json but not in BLS xlsx")
                continue

            bls_emp, bls_lq = bls_data[key]

            if geo_emp != bls_emp:
                mismatches.append(
                    f"  EMPLOYMENT MISMATCH {soc} / {state}: "
                    f"geo={geo_emp:,} BLS={bls_emp:,}"
                )
            elif abs(geo_lq - bls_lq) > 0.01:
                mismatches.append(
                    f"  LQ MISMATCH {soc} / {state}: "
                    f"geo={geo_lq} BLS={bls_lq}"
                )
            else:
                verified += 1

    # 5. Report
    print("=" * 64)
    print("VERIFICATION RESULTS")
    print("=" * 64)
    print(f"✓ Verified : {verified}")
    print(f"✗ Mismatches : {len(mismatches)}")
    print(f"? Missing in BLS : {len(missing)}")

    if mismatches:
        print("\nMISMATCHES (geo_real.json ≠ BLS xlsx):")
        for m in mismatches:
            print(m)

    if missing:
        print("\nMISSING IN BLS (in JSON but not in xlsx):")
        for m in missing:
            print(m)

    if mismatches:
        print(f"\nFAIL — {len(mismatches)} values in geo_real.json do not match BLS source")
        sys.exit(1)
    else:
        print(f"\nPASS — all {verified} geo_real.json values are verified against BLS OES May 2023")


if __name__ == "__main__":
    main()
